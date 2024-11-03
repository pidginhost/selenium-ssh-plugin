import logging
import inspect
from openpyxl import load_workbook
import softest
from typing import List, Optional


class Utils(softest.TestCase):
    """
    Utility class extending softest.TestCase to provide common test functionalities,
    including assertions with logging and Excel data reading.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.logger = self.custom_logger()

    def assert_list_item_text(self, dynamic_list: List, expected_value: str) -> None:
        """
        Asserts that each item's text in the given list matches the expected value.

        This method iterates through a list of elements, retrieves their text,
        and performs a soft assertion to verify that each text matches the expected value.
        It logs relevant information at each step and accumulates any assertion failures.

        Args:
            dynamic_list (List): A list of elements to be checked. Each element should have a `.text` attribute.
            expected_value (str): The expected text value to match against each item's text.

        Raises:
            AssertionError: Aggregated assertion errors if any of the items do not match the expected value.
        """
        self.logger.debug(f"Starting assertion for list items against expected value: '{expected_value}'")

        for index, item in enumerate(dynamic_list, start=1):
            item_text = item.text
            self.logger.debug(f"Item {index}: Retrieved text '{item_text}'")

            # Perform the soft assertion
            self.soft_assert(
                self.assertEquals,
                item_text,
                expected_value,
                f"Item {index}: Expected '{expected_value}', but got '{item_text}'"
            )

            if item_text == expected_value:
                self.logger.info(f"Item {index}: Assertion passed.")
            else:
                self.logger.warning(f"Item {index}: Assertion failed. Expected '{expected_value}', but got '{item_text}'.")

        # After all assertions, check if any failed
        try:
            self.assert_all()
            self.logger.debug("All assertions passed successfully.")
        except AssertionError as e:
            self.logger.error("One or more assertions failed.", exc_info=True)
            raise e

    @staticmethod
    def custom_logger(log_level: int = logging.INFO) -> logging.Logger:
        """
        Creates and returns a custom logger with the specified log level.

        This logger is configured to write logs to 'automation.log' with a specific format.
        It ensures that multiple handlers are not added if the logger already has handlers.

        Args:
            log_level (int): Logging level, default is DEBUG.

        Returns:
            logging.Logger: Configured logger instance.
        """
        # Determine the caller's method name for logger naming
        caller_name = inspect.stack()[1].function
        logger = logging.getLogger(caller_name)
        logger.setLevel(log_level)

        # Avoid adding multiple handlers if the logger already has one
        if not logger.hasHandlers():
            # Create file handler
            fh = logging.FileHandler('automation.log', mode='a')
            # Create formatter and add it to the handler

            formatter = logging.Formatter(
                '%(asctime)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            fh.setFormatter(formatter)
            # Add handler to the logger
            logger.addHandler(fh)

            # Optionally, add a stream handler to output logs to console
            sh = logging.StreamHandler()
            sh.setFormatter(formatter)
            logger.addHandler(sh)

        return logger

    def read_data_from_excel_file(self, excel_file: str, sheet: str) -> Optional[List[List]]:
        """
        Reads data from the specified Excel file and sheet, and returns it as a list of rows.

        Each row is represented as a list of cell values. The method skips the header row
        (assumed to be the first row).

        Args:
            excel_file (str): Path to the Excel file.
            sheet (str): Sheet name to read data from.

        Returns:
            Optional[List[List]]: A list of rows with cell values if successful; otherwise, None.
        """
        self.logger.debug(f"Attempting to read data from Excel file: '{excel_file}', sheet: '{sheet}'")
        datalist = []

        try:
            wb = load_workbook(filename=excel_file, data_only=True)
            sh = wb[sheet]
            row_ct = sh.max_row
            col_ct = sh.max_column

            self.logger.debug(f"Excel sheet '{sheet}' has {row_ct} rows and {col_ct} columns.")

            for i in range(2, row_ct + 1):  # Start from 2 to skip header
                row = [sh.cell(row=i, column=j).value for j in range(1, col_ct + 1)]
                datalist.append(row)
                self.logger.debug(f"Row {i-1}: {row}")

        except FileNotFoundError:
            self.logger.error(f"Error: The file '{excel_file}' was not found.")
            return None
        except KeyError:
            self.logger.error(f"Error: The sheet '{sheet}' does not exist in the workbook '{excel_file}'.")
            return None
        except Exception as e:
            self.logger.error(f"An unexpected error occurred while reading the Excel file: {str(e)}", exc_info=True)
            return None
        else:
            self.logger.info(f"Successfully read {len(datalist)} rows from '{excel_file}' sheet '{sheet}'.")
            return datalist
